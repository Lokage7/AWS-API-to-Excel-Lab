import json
import boto3
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import logging

# Initialize logging
logger = logging.getLogger()
logger.setLevel(logging.INFO)

# Initialize AWS clients
securityhub = boto3.client('securityhub')
s3 = boto3.client('s3')
bedrock = boto3.client('bedrock-runtime')

# Environment variables
S3_BUCKET_NAME = os.environ.get('S3_BUCKET_NAME')
BEDROCK_MODEL_ID = os.environ.get('BEDROCK_MODEL_ID', 'anthropic.claude-3-haiku-20240307-v1:0')
ENABLE_AI_ANALYSIS = os.environ.get('ENABLE_AI_ANALYSIS', 'true').lower() == 'true'

# Compliance framework mappings
COMPLIANCE_MAPPINGS = {
    'SOC 2': {
        'A1.1': ['Access Control', 'Authentication'],
        'A1.2': ['Encryption', 'Data Protection'],
        'A2.1': ['Network Security', 'Firewall'],
        'A3.1': ['Vulnerability Management'],
        'A4.1': ['Monitoring', 'Logging']
    },
    'ISO 27001': {
        'A.9.1': ['Access Control', 'Authentication'],
        'A.10.1': ['Cryptography', 'Encryption'],
        'A.12.1': ['Vulnerability Assessment'],
        'A.13.1': ['Network Security'],
        'A.16.1': ['Incident Management']
    },
    'PCI DSS': {
        '1.1': ['Network Security', 'Firewall'],
        '2.1': ['Encryption', 'Data Protection'],
        '3.1': ['Data Protection'],
        '4.1': ['Monitoring', 'Logging'],
        '7.1': ['Access Control']
    },
    'NIST': {
        'AC-1': ['Access Control'],
        'SC-1': ['Encryption', 'Network Security'],
        'CM-1': ['Configuration Management'],
        'SI-1': ['Monitoring', 'Logging']
    }
}

def invoke_bedrock_for_analysis(findings):
    """Use AWS Bedrock to generate AI-powered analysis of security findings"""
    if not ENABLE_AI_ANALYSIS:
        return {"summary": "AI Analysis Disabled", "recommendations": [], "risk_score": "Medium"}
    
    try:
        # Prepare findings summary for Bedrock
        findings_summary = json.dumps({
            'total_findings': len(findings),
            'critical_findings': len([f for f in findings if f.get('Severity', {}).get('Label') == 'CRITICAL']),
            'high_findings': len([f for f in findings if f.get('Severity', {}).get('Label') == 'HIGH']),
            'medium_findings': len([f for f in findings if f.get('Severity', {}).get('Label') == 'MEDIUM']),
            'low_findings': len([f for f in findings if f.get('Severity', {}).get('Label') == 'LOW']),
            'sample_findings': findings[:5]  # Send first 5 findings for analysis
        })
        
        prompt = f"""
        Analyze these AWS Security Hub findings and provide:
        1. Executive summary (2-3 sentences)
        2. Top 3 prioritized recommendations
        3. Overall risk score (Critical/High/Medium/Low)
        
        Findings data: {findings_summary}
        
        Respond in JSON format:
        {{
            "summary": "Executive summary here",
            "recommendations": ["Recommendation 1", "Recommendation 2", "Recommendation 3"],
            "risk_score": "High"
        }}
        """
        
        # Invoke Bedrock
        response = bedrock.invoke_model(
            ModelId=BEDROCK_MODEL_ID,
            ContentType='application/json',
            Accept='application/json',
            Body=json.dumps({
                "anthropic_version": "bedrock-2023-05-31",
                "max_tokens": 1000,
                "messages": [
                    {
                        "role": "user",
                        "content": prompt
                    }
                ]
            })
        )
        
        response_body = json.loads(response['body'].read().decode('utf-8'))
        ai_analysis = json.loads(response_body['content'][0]['text'])
        
        return ai_analysis
        
    except Exception as e:
        logger.error(f"Bedrock analysis failed: {str(e)}")
        return {"summary": "AI Analysis Failed", "recommendations": [], "risk_score": "Unknown"}

def map_to_compliance_frameworks(finding):
    """Map Security Hub findings to compliance frameworks"""
    mappings = []
    
    # Extract finding details
    title = finding.get('Title', '').lower()
    description = finding.get('Description', '').lower()
    generator_id = finding.get('GeneratorId', '').lower()
    
    # Simple keyword-based mapping
    for framework, controls in COMPLIANCE_MAPPINGS.items():
        for control_id, keywords in controls.items():
            for keyword in keywords:
                if (keyword.lower() in title or 
                    keyword.lower() in description or 
                    keyword.lower() in generator_id):
                    mappings.append({
                        'Framework': framework,
                        'Control': control_id,
                        'Keyword_Matched': keyword
                    })
                    break
    
    return mappings if mappings else [{'Framework': 'Not Mapped', 'Control': 'N/A', 'Keyword_Matched': 'None'}]

def create_enhanced_excel_report(findings, ai_analysis):
    """Create enhanced Excel report with multiple sheets and visualizations"""
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # 1. Executive Summary Sheet
    ws_summary = wb.create_sheet("Executive Summary")
    create_executive_summary(ws_summary, findings, ai_analysis)
    
    # 2. Detailed Findings Sheet
    ws_findings = wb.create_sheet("Detailed Findings")
    create_detailed_findings(ws_findings, findings)
    
    # 3. Compliance Mapping Sheet
    ws_compliance = wb.create_sheet("Compliance Mapping")
    create_compliance_mapping(ws_compliance, findings)
    
    # 4. Dashboard Sheet
    ws_dashboard = wb.create_sheet("Dashboard")
    create_dashboard(ws_dashboard, findings)
    
    # 5. Pivot Analysis Sheet
    ws_pivot = wb.create_sheet("Pivot Analysis")
    create_pivot_analysis(ws_pivot, findings)
    
    return wb

def create_executive_summary(ws, findings, ai_analysis):
    """Create executive summary with AI insights"""
    # Title
    ws['A1'] = "Security Hub Executive Summary"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:E1')
    
    # Date
    ws['A3'] = f"Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A3'].font = Font(bold=True)
    
    # Key Metrics
    ws['A5'] = "Key Metrics"
    ws['A5'].font = Font(size=14, bold=True)
    
    metrics = [
        ("Total Findings", len(findings)),
        ("Critical", len([f for f in findings if f.get('Severity', {}).get('Label') == 'CRITICAL'])),
        ("High", len([f for f in findings if f.get('Severity', {}).get('Label') == 'HIGH'])),
        ("Medium", len([f for f in findings if f.get('Severity', {}).get('Label') == 'MEDIUM'])),
        ("Low", len([f for f in findings if f.get('Severity', {}).get('Label') == 'LOW']))
    ]
    
    for i, (metric, value) in enumerate(metrics, start=6):
        ws[f'A{i}'] = metric
        ws[f'B{i}'] = value
        ws[f'A{i}'].font = Font(bold=True)
    
    # AI Analysis
    ws['A12'] = "AI-Powered Analysis"
    ws['A12'].font = Font(size=14, bold=True)
    
    ws['A13'] = "Summary:"
    ws['B13'] = ai_analysis.get('summary', 'N/A')
    ws['B13'].alignment = Alignment(wrap_text=True)
    
    ws['A14'] = "Risk Score:"
    ws['B14'] = ai_analysis.get('risk_score', 'Unknown')
    
    ws['A15'] = "Top Recommendations:"
    ws['A16'] = "\n".join([f"• {rec}" for rec in ai_analysis.get('recommendations', [])[:3]])
    ws['A16'].alignment = Alignment(wrap_text=True)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 50

def create_detailed_findings(ws, findings):
    """Create detailed findings sheet with all data"""
    headers = [
        "ID", "Title", "Severity", "Criticality", "Compliance Status", 
        "Resource Type", "Region", "First Seen", "Last Seen", 
        "Description", "Remediation", "Generator ID"
    ]
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    
    # Add findings data
    for row, finding in enumerate(findings, 2):
        ws.cell(row=row, column=1, value=finding.get('Id', ''))
        ws.cell(row=row, column=2, value=finding.get('Title', ''))
        ws.cell(row=row, column=3, value=finding.get('Severity', {}).get('Label', ''))
        ws.cell(row=row, column=4, value=finding.get('Severity', {}).get('ProductScore', ''))
        ws.cell(row=row, column=5, value=finding.get('Compliance', {}).get('Status', ''))
        ws.cell(row=row, column=6, value=finding.get('Resources', [{}])[0].get('Type', ''))
        ws.cell(row=row, column=7, value=finding.get('Resources', [{}])[0].get('Region', ''))
        ws.cell(row=row, column=8, value=finding.get('FirstObservedAt', ''))
        ws.cell(row=row, column=9, value=finding.get('LastObservedAt', ''))
        ws.cell(row=row, column=10, value=finding.get('Description', ''))
        ws.cell(row=row, column=11, value=finding.get('Remediation', {}).get('Recommendation', {}).get('Text', ''))
        ws.cell(row=row, column=12, value=finding.get('GeneratorId', ''))
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

def create_compliance_mapping(ws, findings):
    """Create compliance mapping sheet"""
    headers = ["Finding ID", "Title", "Framework", "Control ID", "Keyword Matched"]
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    
    # Add compliance mappings
    row = 2
    for finding in findings:
        mappings = map_to_compliance_frameworks(finding)
        for mapping in mappings:
            ws.cell(row=row, column=1, value=finding.get('Id', ''))
            ws.cell(row=row, column=2, value=finding.get('Title', ''))
            ws.cell(row=row, column=3, value=mapping['Framework'])
            ws.cell(row=row, column=4, value=mapping['Control'])
            ws.cell(row=row, column=5, value=mapping['Keyword_Matched'])
            row += 1
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 25

def create_dashboard(ws, findings):
    """Create dashboard with charts and visualizations"""
    # Title
    ws['A1'] = "Security Dashboard"
    ws['A1'].font = Font(size=16, bold=True)
    
    # Severity distribution data
    severity_data = {
        'Critical': len([f for f in findings if f.get('Severity', {}).get('Label') == 'CRITICAL']),
        'High': len([f for f in findings if f.get('Severity', {}).get('Label') == 'HIGH']),
        'Medium': len([f for f in findings if f.get('Severity', {}).get('Label') == 'MEDIUM']),
        'Low': len([f for f in findings if f.get('Severity', {}).get('Label') == 'LOW'])
    }
    
    # Add data for chart
    ws['A3'] = "Severity"
    ws['B3'] = "Count"
    
    for i, (severity, count) in enumerate(severity_data.items(), start=4):
        ws[f'A{i}'] = severity
        ws[f'B{i}'] = count
    
    # Create bar chart
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Severity Distribution"
    chart.y_axis.title = 'Count'
    chart.x_axis.title = 'Severity'
    
    data = Reference(ws, min_col=2, min_row=3, max_row=7, max_col=2)
    categories = Reference(ws, min_col=1, min_row=4, max_row=7, max_col=1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    
    ws.add_chart(chart, "D3")

def create_pivot_analysis(ws, findings):
    """Create pivot analysis data"""
    headers = ["Severity", "Resource Type", "Region", "Count"]
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    
    # Aggregate data for pivot analysis
    pivot_data = {}
    for finding in findings:
        severity = finding.get('Severity', {}).get('Label', 'Unknown')
        resource_type = finding.get('Resources', [{}])[0].get('Type', 'Unknown')
        region = finding.get('Resources', [{}])[0].get('Region', 'Unknown')
        
        key = (severity, resource_type, region)
        pivot_data[key] = pivot_data.get(key, 0) + 1
    
    # Add pivot data
    row = 2
    for (severity, resource_type, region), count in pivot_data.items():
        ws.cell(row=row, column=1, value=severity)
        ws.cell(row=row, column=2, value=resource_type)
        ws.cell(row=row, column=3, value=region)
        ws.cell(row=row, column=4, value=count)
        row += 1
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

def lambda_handler(event, context):
    """Main Lambda handler"""
    try:
        logger.info("Starting Enhanced Security Hub Excel report generation")
        
        # Get Security Hub findings
        response = securityhub.get_findings(
            MaxResults=1000  # Adjust as needed
        )
        
        findings = response.get('Findings', [])
        logger.info(f"Retrieved {len(findings)} findings")
        
        # Generate AI analysis
        ai_analysis = invoke_bedrock_for_analysis(findings)
        logger.info("AI analysis completed")
        
        # Create enhanced Excel report
        workbook = create_enhanced_excel_report(findings, ai_analysis)
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"security_hub_report_{timestamp}.xlsx"
        
        # Save to S3
        s3_path = f"reports/{filename}"
        workbook.save(f"/tmp/{filename}")
        
        s3.upload_file(
            f"/tmp/{filename}",
            S3_BUCKET_NAME,
            s3_path,
            ExtraArgs={
                'ContentType': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'Metadata': {
                    'ReportType': 'SecurityHubEnhanced',
                    'GeneratedAt': timestamp,
                    'TotalFindings': str(len(findings)),
                    'AIAnalysis': str(ENABLE_AI_ANALYSIS)
                }
            }
        )
        
        logger.info(f"Report uploaded to S3: {s3_path}")
        
        return {
            'statusCode': 200,
            'body': json.dumps({
                'message': 'Enhanced Security Hub Excel report generated successfully',
                'filename': filename,
                's3_path': s3_path,
                'total_findings': len(findings),
                'ai_analysis_enabled': ENABLE_AI_ANALYSIS,
                'sheets_created': ['Executive Summary', 'Detailed Findings', 'Compliance Mapping', 'Dashboard', 'Pivot Analysis']
            })
        }
        
    except Exception as e:
        logger.error(f"Error generating enhanced report: {str(e)}")
        return {
            'statusCode': 500,
            'body': json.dumps({
                'message': 'Error generating Enhanced Security Hub report',
                'error': str(e)
            })
        }
